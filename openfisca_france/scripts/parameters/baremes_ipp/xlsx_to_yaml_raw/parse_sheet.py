# -*- coding: utf-8 -*-

import logging
import re
import traceback
import datetime
import collections
import sys
import os

import yaml
from biryani import strings
from openpyxl.utils import range_boundaries

from parse_cell import read_cell

log = logging.getLogger('xlsx_parser')

# AAD=Age d'annulation de la décôte
aad_re = re.compile(ur'AAD(\s+[-+]\s+\d+\s+ans?)?$')


def range_boundaries_zero_indexes(range_string):
    column_low, row_low, column_high, row_high = range_boundaries(range_string)
    return column_low - 1, row_low - 1, column_high, row_high


def extract_merged_cells(sheet):
    merged_cells_tree = {}
    for range_string in sheet.merged_cell_ranges:
        column_low, row_low, column_high, row_high = range_boundaries_zero_indexes(range_string)
        for row_index in range(row_low, row_high):
            cell_coordinates_by_merged_column_index = merged_cells_tree.setdefault(
                row_index, {})
            for column_index in range(column_low, column_high):
                cell_coordinates_by_merged_column_index[column_index] = (row_low, column_low)
    return merged_cells_tree


def parse_sheet(book, sheet_name, sheet_title_by_name, book_yaml_dir_encoded):
    log.info(u'  Parsing sheet {}.'.format(sheet_name))
    sheet = book.get_sheet_by_name(sheet_name)
    sheet_error = None

    try:
        merged_cells_tree = extract_merged_cells(sheet)

        if sheet_name.startswith((u'Sommaire', u'Outline')):
            pass
        elif sheet_name.startswith(u'Abréviation'):
            pass
        else:
            sheet_node, yaml_file_path_encoded = parse_content_sheet(book, sheet, sheet_name, merged_cells_tree, book_yaml_dir_encoded)

            with open(yaml_file_path_encoded, 'w') as yaml_file:
                yaml.dump(sheet_node, yaml_file, allow_unicode = True, default_flow_style = False, indent = 2, width = 120)
    except:
        message = u'An exception occurred when parsing sheet "{}".'.format(sheet_name)
        log.exception(u'    {}'.format(message))
        sheet_error = literal_unicode(u'\n\n'.join(
            fragment
            for fragment in (
                unicode(sheet_error) if sheet_error is not None else None,
                message,
                traceback.format_exc().decode('utf-8'),
                )
            if fragment
            ))

    return sheet_title_by_name, sheet_error

def is_a_date(first_cell_value):
    if isinstance(first_cell_value, datetime.date):
        return first_cell_value
    elif isinstance(first_cell_value, int):
        if 1914 < first_cell_value and first_cell_value < 2020:
            return datetime.date(first_cell_value, 1, 1)
        return None
    elif isinstance(first_cell_value, basestring) and aad_re.match(first_cell_value):
        return first_cell_value
    else:
        return None

def parse_content_sheet(book, sheet, sheet_name, merged_cells_tree, book_yaml_dir_encoded):
    descriptions_rows = []
    labels_rows = []
    notes_rows = []
    state = 'taxipp_names'
    taxipp_names_row = None
    values_rows = []
    for row_index in range(sheet.max_row):
        columns_count = sheet.max_column
        if state == 'taxipp_names':
            taxipp_names_row = [
                (taxipp_name or u'').strip()
                for taxipp_name in (
                    read_cell(book, sheet, merged_cells_tree, row_index, column_index)
                    for column_index in range(columns_count)
                    )
                ]
            state = 'labels'
            #log.info(taxipp_names_row)
            if all(
                    not taxipp_name
                    for taxipp_name in taxipp_names_row
                    ):
                # The first row is empty => This sheet doesn't contain TaxIPP names.
                continue
            # When any TaxIPP name is in lowercase, assume that this row is really the TaxIPP names row.
            if any(
                    taxipp_name and taxipp_name[0].islower()
                    for taxipp_name in taxipp_names_row
                    ):
                continue
            else:
                log.info(u'    Sheet "{}" has no row for TaxIPP names.'.format(sheet_name))
                taxipp_names_row = []
        if state == 'labels':
            first_cell_value = read_cell(book, sheet, merged_cells_tree, row_index, 0)
            if not is_a_date(first_cell_value):
                # First cell of row is not a the first cell of a row of values => Assume it is a label.
                labels_row = []
                for column_index in range(columns_count):
                    label = read_cell(book, sheet, merged_cells_tree, row_index, column_index)
                    label = label or u''
                    label = unicode(label)  # Some column labels are parsed as datetime.date
                    label = u' '.join(label.split()).strip()
                    labels_row.append(label)
                labels_rows.append(labels_row)
                continue
            #log.info('state=value, row_index={}'.format(row_index))
            state = 'values'
        if state == 'values':
            first_cell_value = read_cell(book, sheet, merged_cells_tree, row_index, 0)
            if is_a_date(first_cell_value):
                first_cell_value = is_a_date(first_cell_value)
                # First cell of row is a valid date or year.
                values_row = [
                    value.strip() if isinstance(value, basestring) else value
                    for value in (
                        read_cell(book, sheet, merged_cells_tree, row_index, column_index, empty_white_value = u'nc')
                        for column_index in range(columns_count)
                        )
                    ]
                if isinstance(first_cell_value, datetime.date):
                    assert first_cell_value.year < 2601, 'Invalid date {} in {} at row {}'.format(first_cell_value, sheet_name, row_index + 1)
                    values_rows.append(values_row)
                    continue
                if isinstance(first_cell_value, basestring) and aad_re.match(first_cell_value) is not None:
                    values_rows.append(values_row)
                    continue
                if all(value in (None, u'', u'nc') for value in values_row):
                    # If first cell is empty and all other cells in line are also empty, ignore this
                    # line.
                    continue
                # First cell has no date and other cells in row are not empty => Assume it is a note.
            #log.info('state=notes, row_index={}'.format(row_index))
            break

    sheet_node = collections.OrderedDict()

    sheet_node[u'Titre court'] = sheet_name

    labels = []
    for labels_row in labels_rows:
        for column_index, label in enumerate(labels_row):
            if label is None:
                continue
            label = label.strip()
            if not label:
                continue
            while column_index >= len(labels):
                labels.append([])
            column_labels = labels[column_index]
            if not column_labels or column_labels[-1] != label:
                column_labels.append(label)
    labels = [
        (tuple(
            label_stripped
            for label_stripped in (
                (label or u'').strip()
                for label in column_labels1
                )
            if label_stripped
            ) if column_labels1 else None) or (u'Colonne sans titre',)
        for index, column_labels1 in enumerate(labels, 1)
        ]
    assert labels

    taxipp_name_by_column_labels = collections.OrderedDict()
    for column_labels, taxipp_name in zip(labels, taxipp_names_row):
        if not taxipp_name:
            continue
        taxipp_name_by_column_label = taxipp_name_by_column_labels
        for column_label in column_labels[:-1]:
            taxipp_name_by_column_label = taxipp_name_by_column_label.setdefault(column_label,
                collections.OrderedDict())
        taxipp_name_by_column_label[column_labels[-1]] = taxipp_name
    if taxipp_name_by_column_labels:
        sheet_node[u'Noms TaxIPP'] = taxipp_name_by_column_labels

    sheet_values = []
    for value_row in values_rows:
        cell_by_column_labels = collections.OrderedDict()
        for column_labels, cell in zip(labels, value_row):
            if cell is None or cell == '':
                continue
            cell_by_column_label = cell_by_column_labels
            for column_label in column_labels[:-1]:
                cell_by_column_label = cell_by_column_label.setdefault(column_label,
                    collections.OrderedDict())
            # Merge (amount, unit) couples to a string to simplify YAML.
            if isinstance(cell, datetime.date):
                cell = str(cell)  # TODO : remove
            if isinstance(cell, tuple):
                cell = unicode(cell[0]) + u' ' + unicode(cell[1])
            if isinstance(cell, basestring) and u'\n' in cell:
                cell = literal_unicode(cell)
            cell_by_column_label[column_labels[-1]] = cell
        sheet_values.append(cell_by_column_labels)
    if sheet_values:
        sheet_node[u'Valeurs'] = sheet_values

    file_system_encoding = sys.getfilesystemencoding()
    yaml_file_path_encoded = os.path.join(
        book_yaml_dir_encoded,
        (strings.slugify(sheet_name) + u'.yaml').encode(file_system_encoding),
        )

    return sheet_node, yaml_file_path_encoded


def get_hyperlink(sheet, row_index, column_index):
    cell = sheet.cell(row=row_index+1, column=column_index+1)
    return cell.hyperlink


class literal_unicode(unicode):
    pass
